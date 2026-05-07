"""
LL Árváltozás Detektor
======================
Összehasonlít két LL Excel árlistát és kilistázza a változásokat:
  - Megváltozott árak (normál és akciós)
  - Megváltozott kiszerelések
  - Új termékek
  - Törölt termékek

Használat:
  python arvaltozas.py <regi.xlsx> <uj.xlsx>

Példa:
  python arvaltozas.py LL/regi_arlista.xlsx LL/uj_arlista.xlsx
"""

import sys
from datetime import datetime
from openpyxl import load_workbook


# ── Excel beolvasás ────────────────────────────────────────────────────────────

def beolvas(fajl_ut: str) -> list[dict]:
    """
    Beolvas egy .xlsx fájlt, visszaad egy listát szótárakkal.
    Mindkét LL formátumot kezeli:
      - Régi (makró): Első_sor, Második_sor, Harmadik_sor, Cikkszám, Kiszerelés, Ár, Akciós_ár
      - Új (sablon):  Megnevezés, Cikkszám, Kiszerelés, Ár, Akciós_ár
    """
    wb = load_workbook(fajl_ut, data_only=True)
    ws = wb.active

    sorok = list(ws.iter_rows(values_only=True))
    if not sorok:
        return []

    # Első sor = fejléc, None cellákat kihagyjuk
    fejlec = [str(c).strip() if c is not None else f"oszlop_{i}"
              for i, c in enumerate(sorok[0])]

    termekek = []
    for sor in sorok[1:]:
        d = {fejlec[i]: (str(v).strip() if v is not None else "")
             for i, v in enumerate(sor)}
        # Teljesen üres sor kihagyása
        if all(v == "" for v in d.values()):
            continue
        termekek.append(d)

    return termekek


# ── Segédfüggvények ────────────────────────────────────────────────────────────

def normalize_ar(val) -> str:
    """Árból csak a számjegyeket tartja meg: '8 990', '8.990', 8990 → '8990'"""
    if val is None or val == "":
        return ""
    return "".join(c for c in str(val) if c.isdigit())


def megnevezes(sor: dict) -> str:
    """Emberi olvasásra szánt terméknév."""
    if sor.get("Megnevezés"):
        return sor["Megnevezés"]
    reszek = [sor.get("Első_sor", ""), sor.get("Második_sor", ""), sor.get("Harmadik_sor", "")]
    return " ".join(r for r in reszek if r).strip()


def cikk_kulcs(sor: dict) -> str:
    """
    Rendezési és csoportosítási kulcs egy sorhoz.
    Elsőbbség: Cikkszám → Megnevezés → összefűzött sorok.
    Cikkszám esetén numerikusan rendez (pl. '500156' < '1000000').
    """
    cikk = sor.get("Cikkszám", "").strip()
    if cikk:
        # numerikus rendezéshez nullával töltjük fel
        return cikk.zfill(20)
    megn = sor.get("Megnevezés", "").strip().lower()
    if megn:
        return megn
    e1 = sor.get("Első_sor", "").strip().lower()
    e2 = sor.get("Második_sor", "").strip().lower()
    e3 = sor.get("Harmadik_sor", "").strip().lower()
    return " ".join(x for x in [e1, e2, e3] if x)


def csoportosit(sorok: list[dict]) -> dict[str, list[dict]]:
    """
    Cikkszám (vagy névkulcs) szerint csoportosítja a sorokat.
    Visszaad egy rendezett dict-et: kulcs → [sor1, sor2, sor3, ...]
    Figyelmeztet, ha egy csoportban több mint 3 sor van (adatminőségi hiba).
    """
    csoportok: dict[str, list[dict]] = {}
    for sor in sorok:
        k = cikk_kulcs(sor)
        if not k:
            continue  # teljesen azonosíthatatlan sort kihagyjuk
        csoportok.setdefault(k, []).append(sor)

    for k, csoport in csoportok.items():
        if len(csoport) > 3:
            print(f"  [FIGYELEM] '{k}' kulcshoz {len(csoport)} sor tartozik (várható: max 3)")

    return csoportok


# ── Összehasonlítás ────────────────────────────────────────────────────────────

def sorok_osszehasonlitasa(regi_csoport: list[dict], uj_csoport: list[dict],
                           eredmeny: dict) -> None:
    """
    Két termék sorcsoportját hasonlítja össze pozíció alapján.
    Mindkét csoportban max 3 sor van; row[i] régi vs row[i] uj.
    """
    min_len = min(len(regi_csoport), len(uj_csoport))

    for i in range(min_len):
        regi_sor = regi_csoport[i]
        uj_sor   = uj_csoport[i]

        regi_ar    = normalize_ar(regi_sor.get("Ár"))
        uj_ar      = normalize_ar(uj_sor.get("Ár"))
        regi_akcio = normalize_ar(regi_sor.get("Akciós_ár"))
        uj_akcio   = normalize_ar(uj_sor.get("Akciós_ár"))
        regi_kisz  = regi_sor.get("Kiszerelés", "").strip()
        uj_kisz    = uj_sor.get("Kiszerelés", "").strip()

        uj_van_ar   = bool(uj_ar or uj_akcio)
        regi_van_ar = bool(regi_ar or regi_akcio)

        # Ha a sor teljesen üressé vált → törölt, nem vizsgáljuk tovább
        if regi_van_ar and not uj_van_ar:
            eredmeny["torolt_termek"].append(regi_sor)
            continue

        if regi_ar != uj_ar and uj_ar:
            eredmeny["ar_valtozas"].append((regi_sor, uj_sor))

        if regi_akcio != uj_akcio and (regi_akcio or uj_akcio):
            eredmeny["akscios_ar_valtozas"].append((regi_sor, uj_sor))

        if regi_kisz != uj_kisz and uj_kisz:
            eredmeny["kiszeres_valtozas"].append((regi_sor, uj_sor))

    # A régi csoportban maradt, az újban nincs → törölt sorok
    for sor in regi_csoport[min_len:]:
        if normalize_ar(sor.get("Ár")) or normalize_ar(sor.get("Akciós_ár")):
            eredmeny["torolt_termek"].append(sor)

    # Az új csoportban van, a régiben nincs → új sorok ennél a terméknél
    for sor in uj_csoport[min_len:]:
        if normalize_ar(sor.get("Ár")) or normalize_ar(sor.get("Akciós_ár")):
            eredmeny["uj_termek"].append(sor)


def osszehasonlit(regi: list[dict], uj: list[dict]) -> dict[str, list]:
    """
    Összehasonlítja a két árlistát.
    1. Mindkét listát Cikkszám szerint csoportosítja és rendezi.
    2. Egyező kulcsú csoportokat pozíció alapján hasonlítja össze.
    3. Csak az újban lévő csoport → új termék.
    4. Csak a régiben lévő csoport → törölt termék.
    """
    eredmeny = {
        "ar_valtozas":         [],
        "akscios_ar_valtozas": [],
        "kiszeres_valtozas":   [],
        "uj_termek":           [],
        "torolt_termek":       [],
    }

    regi_csop = csoportosit(regi)
    uj_csop   = csoportosit(uj)

    regi_kulcsok = set(regi_csop)
    uj_kulcsok   = set(uj_csop)

    # Mindkét fájlban meglévő termékek összehasonlítása
    for kulcs in sorted(regi_kulcsok & uj_kulcsok):
        sorok_osszehasonlitasa(regi_csop[kulcs], uj_csop[kulcs], eredmeny)

    # Csak az újban lévő termékek → új termékek
    for kulcs in sorted(uj_kulcsok - regi_kulcsok):
        for sor in uj_csop[kulcs]:
            if normalize_ar(sor.get("Ár")) or normalize_ar(sor.get("Akciós_ár")):
                eredmeny["uj_termek"].append(sor)

    # Csak a régiben lévő termékek → törölt termékek
    for kulcs in sorted(regi_kulcsok - uj_kulcsok):
        for sor in regi_csop[kulcs]:
            if normalize_ar(sor.get("Ár")) or normalize_ar(sor.get("Akciós_ár")):
                eredmeny["torolt_termek"].append(sor)

    return eredmeny


# ── Riport kiírás ──────────────────────────────────────────────────────────────

def format_ar_delta(regi_ar: str, uj_ar: str) -> str:
    try:
        delta = int(uj_ar) - int(regi_ar)
        return f"({'+' if delta > 0 else ''}{delta} Ft)"
    except ValueError:
        return ""


def riport_kiir(eredmeny: dict, regi_fajl: str, uj_fajl: str):
    osszes = sum(len(v) for v in eredmeny.values())

    if osszes == 0:
        print("\n==> Nincs változás a két fájl között.\n")
        return

    print(f"\n{'='*60}")
    print(f"  ÁRVÁLTOZÁS RIPORT — {osszes} változás")
    print(f"  Régi: {regi_fajl}")
    print(f"  Új:   {uj_fajl}")
    print(f"{'='*60}")

    if eredmeny["ar_valtozas"]:
        print(f"\n  ÁR VÁLTOZOTT ({len(eredmeny['ar_valtozas'])} db):")
        print(f"  {'-'*56}")
        for regi_sor, uj_sor in eredmeny["ar_valtozas"]:
            nev   = megnevezes(uj_sor)
            delta = format_ar_delta(regi_sor.get("Ár",""), uj_sor.get("Ár",""))
            print(f"  ! {nev:<38} {regi_sor.get('Ár','')} -> {uj_sor.get('Ár','')} Ft  {delta}")

    if eredmeny["akscios_ar_valtozas"]:
        print(f"\n  AKCIÓS ÁR VÁLTOZOTT ({len(eredmeny['akscios_ar_valtozas'])} db):")
        print(f"  {'-'*56}")
        for regi_sor, uj_sor in eredmeny["akscios_ar_valtozas"]:
            nev    = megnevezes(uj_sor)
            r_a    = regi_sor.get("Akciós_ár","") or "—"
            u_a    = uj_sor.get("Akciós_ár","")   or "—"
            print(f"  * {nev:<38} {r_a} -> {u_a} Ft")

    if eredmeny["kiszeres_valtozas"]:
        print(f"\n  KISZERELÉS VÁLTOZOTT ({len(eredmeny['kiszeres_valtozas'])} db):")
        print(f"  {'-'*56}")
        for regi_sor, uj_sor in eredmeny["kiszeres_valtozas"]:
            nev = megnevezes(uj_sor)
            print(f"  ~ {nev:<38} [{regi_sor.get('Kiszerelés','')}] -> [{uj_sor.get('Kiszerelés','')}]")

    if eredmeny["uj_termek"]:
        print(f"\n  ÚJ TERMÉKEK ({len(eredmeny['uj_termek'])} db):")
        print(f"  {'-'*56}")
        for sor in eredmeny["uj_termek"]:
            print(f"  + {megnevezes(sor):<38} {sor.get('Ár','')} Ft")

    if eredmeny["torolt_termek"]:
        print(f"\n  TÖRÖLT TERMÉKEK ({len(eredmeny['torolt_termek'])} db):")
        print(f"  {'-'*56}")
        for sor in eredmeny["torolt_termek"]:
            print(f"  - {megnevezes(sor):<38} volt: {sor.get('Ár','')} Ft")

    print(f"\n{'─'*60}")
    print("  ÖSSZESÍTŐ:")
    kategoria_nevek = {
        "ar_valtozas":         "Ár változott",
        "akscios_ar_valtozas": "Akciós ár változott",
        "kiszeres_valtozas":   "Kiszerelés változott",
        "uj_termek":           "Új termékek",
        "torolt_termek":       "Törölt termékek",
    }
    for key, sorok in eredmeny.items():
        if sorok:
            print(f"  {kategoria_nevek[key]:<30} {len(sorok)} db")
    print()


def riport_fajlba(eredmeny: dict, regi_fajl: str, uj_fajl: str):
    datum   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    kimenet = f"riport_{datum}.txt"

    osszes = sum(len(v) for v in eredmeny.values())

    with open(kimenet, "w", encoding="utf-8") as f:
        f.write("LL ÁRVÁLTOZÁS RIPORT\n")
        f.write(f"Dátum:     {datum.replace('_', ' ')}\n")
        f.write(f"Régi fájl: {regi_fajl}\n")
        f.write(f"Új fájl:   {uj_fajl}\n")
        f.write("="*60 + "\n\n")

        if osszes == 0:
            f.write("Nincs változás.\n")
        else:
            kategoria_nevek = {
                "ar_valtozas":         "ÁR VÁLTOZOTT",
                "akscios_ar_valtozas": "AKCIÓS ÁR VÁLTOZOTT",
                "kiszeres_valtozas":   "KISZERELÉS VÁLTOZOTT",
                "uj_termek":           "ÚJ TERMÉKEK",
                "torolt_termek":       "TÖRÖLT TERMÉKEK",
            }
            for key, sorok in eredmeny.items():
                if not sorok:
                    continue
                f.write(f"[ {kategoria_nevek[key]} — {len(sorok)} db ]\n")
                for elem in sorok:
                    if isinstance(elem, tuple):
                        regi_sor, uj_sor = elem
                        nev = megnevezes(uj_sor)
                        if key == "ar_valtozas":
                            delta = format_ar_delta(regi_sor.get("Ár",""), uj_sor.get("Ár",""))
                            f.write(f"  ! {nev} — {regi_sor.get('Ár','')} -> {uj_sor.get('Ár','')} Ft {delta}\n")
                        elif key == "akscios_ar_valtozas":
                            f.write(f"  * {nev} — {regi_sor.get('Akciós_ár','') or '—'} -> {uj_sor.get('Akciós_ár','') or '—'} Ft\n")
                        elif key == "kiszeres_valtozas":
                            f.write(f"  ~ {nev} — [{regi_sor.get('Kiszerelés','')}] -> [{uj_sor.get('Kiszerelés','')}]\n")
                    else:
                        prefix = "+" if key == "uj_termek" else "-"
                        f.write(f"  {prefix} {megnevezes(elem)} — {elem.get('Ár','')} Ft\n")
                f.write("\n")

    print(f"Riport mentve: {kimenet}")


# ── Belépési pont ──────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Használat: python arvaltozas.py <regi.xlsx> <uj.xlsx>")
        print("Példa:     python arvaltozas.py regi_arlista.xlsx uj_arlista.xlsx")
        sys.exit(1)

    regi_fajl = sys.argv[1]
    uj_fajl   = sys.argv[2]

    print(f"\nRégi fájl beolvasása: {regi_fajl}")
    regi = beolvas(regi_fajl)
    print(f"  -> {len(regi)} termék")

    print(f"Új fájl beolvasása:   {uj_fajl}")
    uj = beolvas(uj_fajl)
    print(f"  -> {len(uj)} termék")

    eredmeny = osszehasonlit(regi, uj)
    riport_kiir(eredmeny, regi_fajl, uj_fajl)
    riport_fajlba(eredmeny, regi_fajl, uj_fajl)


if __name__ == "__main__":
    main()
